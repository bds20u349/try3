from name_iops_avg import Reading_the_necessary_information
import re
from openpyxl import load_workbook

def Output_from_out_iostat(file_name):
    massiv = []
    file = open(file_name, 'r')
    a = True 
    row = 2
    #пропускаем строчки, нет интересуещих нас значений
    file.readline()
    file.readline()
    line = file.readline()

    try:
        while a:
            #считываем информацию о каждом устройстве и записываем в excel
            while (line!= '\n') and (line):
                massiv.append(line.split())
                line = file.readline()
            if not line:
                a = False
            #пропускаем строчки, нет интересуещих нас значений
            file.readline()
            file.readline()
            line = file.readline()
    finally:
        file.close()
        return massiv

name_of_the_folder_out_top = input('введите путь к файлу out_iostat: ').replace('\\', '/')
name_of_the_out_top = input('введите имя файла out_iostat: ')
name_of_the_folder_out_top += '/' + name_of_the_out_top
name_of_the_excel = input('введите полный путь к файлу excel: ').replace('\\', '/')

wb = load_workbook(name_of_the_excel)#получаем доступ к excel файлу
if name_of_the_out_top in wb.sheetnames: #если лист с таки именем уже существует, то удаляем его и создаём снова, тем самым отчищаем его от прошлых значений
    wb.remove(wb[name_of_the_out_top])
ws = wb.create_sheet(name_of_the_out_top) #создаём новый лист

massiv = Output_from_out_iostat(name_of_the_folder_out_top)
for i in massiv:
    ws.append(i)

wb.save(name_of_the_excel) #сохраняем изменения в лист excel
wb.close() #закрываем лист excel
