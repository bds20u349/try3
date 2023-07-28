from name_iops_avg import Output
from name_iops_avg import Checking_the_file_name
from name_iops_avg import File_name_sorting_function
from export_excel import ExportExcel
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Series, Reference

import os

#-------------------Открываем файл и считываем оттуда пути к папкам с тестами и папке с excel----------------------------------------------------
config_file = open('Config.txt', 'r') #открываем файл, в торый записаны пути к каталогу с тестами и к котологу с excel файлом
name_of_the_folder_with_tests = config_file.readline()[13:-1] #первые 13 символов - это "folder name: ", а последний знак переноса "\n" файлом
name_of_the_folder_with_excel = config_file.readline()[11:-1] #первые 11 символов - это "file name: ", а последний знак переноса "\n"
config_file.close() #закрываем файл
#-----------------------------------------------------------------------------------------------------------------------------------------------



#---------Считываем имя интересуещей нас папки с тестами и имя интресуещего на excel файла. Сотавляем общий путь к excel файлу и файлам с тестами--------------
name_of_the_folder_with_tests = name_of_the_folder_with_tests.replace('\\', '/') #заменяем системные знаки \ на / в пути к папке с тестами
name_of_the_folder_with_excel = name_of_the_folder_with_excel.replace('\\', '/') #заменяем системные знаки \ на / в пути к папке с excel
folder_name = input('введите имя папки с тестами: ') #rw6d4-16kstrip1800slong
excel_name = input('введите имя файла excel: ') + '.xlsx' #test
name_of_the_folder_with_tests = name_of_the_folder_with_tests + '/' + folder_name #Получаем общий путь внутрь каталога с тестами
name_of_the_folder_with_excel = name_of_the_folder_with_excel + '/' + excel_name #Получаем общий путь к файлу excel
catalog_with_tests = os.listdir(name_of_the_folder_with_tests) #получаем список названий всех файлов внутри каталога
catalog_with_tests.sort(key=File_name_sorting_function)
#--------------------------------------------------------------------------------------------------------------------------------------------------------------



wb = load_workbook(name_of_the_folder_with_excel)#получаем доступ к excel файлу
if folder_name in wb.sheetnames: #если лист с таки именем уже существует, то удаляем его и создаём снова, тем самым отчищаем его от прошлых значений
    wb.remove(wb[folder_name])
ws = wb.create_sheet(folder_name) #создаём новый лист

for file_name in catalog_with_tests: #берём по очереди все файлы с тестами
    if Checking_the_file_name(file_name): #проверка на соответствие имени файла шаблону вида "fio-*-*.log"
        name, iops, avg = Output(name_of_the_folder_with_tests + '/' + file_name) #Получение из файла с результатом теста имени тестируемого диска, отправленных iops и результата evg
        max_row = ExportExcel(ws, name, iops, avg) #записываем данные в excel

#-------------------------------------------------------------------------------
row, col = 1, 1
while(ws.cell(row = row, column = col).value != None):
    ch = ScatterChart()
    xvalues = Reference(ws, min_col=col,  max_col=col, min_row=3, max_row=max_row)
    values = Reference(ws, min_col=col+1, max_col=col+1, min_row=2, max_row=max_row)
    series = Series(values, xvalues, title_from_data=True)
    series.marker.size = 5
    series.marker.symbol = "diamond"
    # добавляем данные в объект диаграммы
    ch.series.append(series)
    ch.title = folder_name + '(' + str(ws.cell(row = 1, column = col).value) + ')'
    ch.x_axis.title = 'I/Os per Second'
    ch.y_axis.title = 'Response (msec)'

    ws.add_chart(ch, "A" + str(max_row+col*7-5))
    col += 2

wb.save(name_of_the_folder_with_excel) #сохраняем изменения в лист excel
wb.close() #закрываем лист excel
