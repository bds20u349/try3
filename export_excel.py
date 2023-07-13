from openpyxl import load_workbook

def ExportExcel(ws, name, iops, avg):
    row = 1 #начинаем с верхней, левой ячейки
    col = 1
    a = True

    while (a): #цикл продолжает работу, пока данные не будут записаны (а - флаг записи данных)
        cell = ws.cell(row = row, column = col) #получаем значение ячейки с именем диска
        if (cell.value != name) and (cell.value != None): #если не пустая и имя в ячейки не совпадает с заданным именем, то
            col += 2 #перескакиваем на следующую ячейку с именем
        else:
            if (cell.value == None): #Если имени диска ещё нет
                ws.cell(row = row, column = col).value = name 
                ws.cell(row = row+1, column = col).value = 'iops'
                ws.cell(row = row+1, column = col+1).value = 'avg'

            while(ws.cell(row = row, column = col).value != None): #ищем первую свободную пару ячеек iops-evg в столбце
                row += 1                

            ws.cell(row = row, column = col).value = iops
            ws.cell(row = row, column = col+1).value = avg
            a = False
