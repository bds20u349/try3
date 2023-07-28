from name_iops_avg import Reading_the_necessary_information
import re
from openpyxl import load_workbook

def Output_from_out_top(file_name):
    file = open(file_name, 'r')
    massiv = []
    a = True #флаг нахождения последнего, интересуещего нас значения
    clovar = {'time': 0,'load avg 1min': 0,'load avg 5min': 0,'load avg 15min': 0,'сумма всех %CPU': 0,'IO wait': 0}      
    try:
        while a:
            clovar.fromkeys(clovar, 0)
            summa_spu = 0 #сумма CPU
            line = file.readline() #считали строку
            m = re.search('\d\d[:]\d\d[:]\d\d', line) #находим время
            clovar['time'] = m.group(0)
            m = re.search("[:]\s(.+)[,]\s(.+)[,]\s(.+)", line) #находим значения load average
            clovar['load avg 1min'], clovar['load avg 5min'], clovar['load avg 15min'] = m.group(1), m.group(2), m.group(3)
            #пропускаем строчки, нет интересуещих нас значений
            line = file.readline()
            line = file.readline()
            m = re.search("id,\s(.+)\swa", line) #находим значение IO wait
            clovar['IO wait'] = m.group(1)
            #пропускаем строчки, нет интересуещих нас значений
            line = file.readline()
            m = re.search("([\w.]+)\sbuff/cache", line) #находим значение buff/cache
            clovar['buff/cache'] = m.group(1)
            #пропускаем строчки, нет интересуещих нас значений
            line = file.readline()
            line = file.readline()
            line = file.readline()
            line = file.readline()
            #считываем информацию о работе каждого PID
            while(line!= '\n') and (line):
                stroka = line.split()
                pid, cpu = stroka[0], stroka[8]
                summa_spu += float(cpu)
                clovar['PID № ' + str(pid)] = cpu
                line = file.readline()
            clovar['сумма всех %CPU'] = summa_spu
            massiv.append(clovar.copy())
            if not line: #пока не конец файла
                a = False
    finally:
       file.close()
       return massiv

name_of_the_folder_out_top = input('введите путь к файлу out_top: ').replace('\\', '/')
name_of_the_out_top = input('введите имя файла out_top: ')
name_of_the_folder_out_top += '/' + name_of_the_out_top
name_of_the_excel = input('введите полный путь к файлу excel: ').replace('\\', '/')

wb = load_workbook(name_of_the_excel)#получаем доступ к excel файлу
if name_of_the_out_top in wb.sheetnames: #если лист с таки именем уже существует, то удаляем его и создаём снова, тем самым отчищаем его от прошлых значений
    wb.remove(wb[name_of_the_out_top])
ws = wb.create_sheet(name_of_the_out_top) #создаём новый лист

massiv = Output_from_out_top(name_of_the_folder_out_top)
ws.append(list(massiv[-1]))
for i in massiv:
    ws.append(list(i.values()))
wb.save(name_of_the_excel) #сохраняем изменения в лист excel
wb.close() #закрываем лист excel
