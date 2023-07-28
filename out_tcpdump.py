from name_iops_avg import Reading_the_necessary_information
import re
from openpyxl import load_workbook

def Output_from_out_tcpdump(file_name):
    file = open(file_name, 'r')
    massiv = []
    a = True #флаг нахождения последнего, интересуещего нас значения
    clovar = {'time': 0,'Flags': 0,'length': 0}      
    try:
        while a:
            clovar.fromkeys(clovar, 0)
            line = file.readline() #считали строку
            m = re.search('(\d\d[:]\d\d[:]\d\d[.]\d+)\s', line) #находим время
            clovar['time'] = m.group(1)
            m = re.search('Flags\s([^,]+),', line) #находим время
            clovar['Flags'] = m.group(1)
            m = re.search('length\s(\d+)', line) #находим время
            clovar['length'] = m.group(1)
            massiv.append(clovar.copy())
            if not line: #пока не конец файла
                a = False
    finally:
       file.close()
       return massiv

name_of_the_folder_out_top = input('введите путь к файлу out_tcpdump: ').replace('\\', '/')
name_of_the_out_top = input('введите имя файла out_tcpdump: ')
name_of_the_folder_out_top += '/' + name_of_the_out_top
name_of_the_excel = input('введите полный путь к файлу excel: ').replace('\\', '/')

wb = load_workbook(name_of_the_excel)#получаем доступ к excel файлу
if name_of_the_out_top in wb.sheetnames: #если лист с таки именем уже существует, то удаляем его и создаём снова, тем самым отчищаем его от прошлых значений
    wb.remove(wb[name_of_the_out_top])
ws = wb.create_sheet(name_of_the_out_top) #создаём новый лист

massiv = Output_from_out_tcpdump(name_of_the_folder_out_top)
ws.append(list(massiv[-1]))
for i in massiv:
    ws.append(list(i.values()))
wb.save(name_of_the_excel) #сохраняем изменения в лист excel
wb.close() #закрываем лист excel
